# -*- coding: utf-8 -*-
"""產生填寫式 Excel 報名表：每位學生一列，項目欄用下拉選單
班導打開 → 選項目 → 存檔 → 上傳/交回 → 用 import_data.py regs 匯入"""
import sqlite3
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from rules import DB_PATH, load_config

BASE = Path(__file__).parent
TEMPLATE_DIR = BASE / "templates"
TEMPLATE_DIR.mkdir(exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Microsoft JhengHei")
NORMAL_FONT = Font(name="Microsoft JhengHei")
INSTR_FILL = PatternFill("solid", fgColor="FFF2CC")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fetch():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT student_id, name, grade, class_no, seat_no, gender, bib_number "
                "FROM students ORDER BY grade, class_no, seat_no")
    students = cur.fetchall()
    cur.execute("SELECT code, name, category, gender, grade_limit FROM events "
                "ORDER BY category, name")
    events = cur.fetchall()
    conn.close()
    return students, events


def _apply_header(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def make_form(grade_filter=None, class_filter=None, out_name=None):
    cfg = load_config()
    meet = cfg["meet_info"]
    rule = cfg["track_field_rules"]["options"][cfg["track_field_rules"]["active_option"]]
    students, events = _fetch()

    if grade_filter:
        students = [s for s in students if s[2] == grade_filter]
    if class_filter:
        students = [s for s in students if s[3] == class_filter]

    if not students:
        print(f"[X] 找不到符合條件的學生 (grade={grade_filter}, class={class_filter})")
        return None

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "使用說明"
    ws1.column_dimensions["A"].width = 90
    ws1["A1"] = f"{meet['year']} {meet['name']}  報名表使用說明"
    ws1["A1"].font = Font(name="Microsoft JhengHei", bold=True, size=14)
    ws1["A2"] = f"目前規則：{cfg['track_field_rules']['active_option']}"
    ws1["A2"].font = Font(name="Microsoft JhengHei", bold=True, color="C00000")
    ws1["A3"] = (f"個人上限 {rule['individual_max']} 項"
                 f"（田賽 {rule['field_max']}、徑賽 {rule['track_max']}），"
                 f"接力上限 {rule['relay_max']} 項")
    instructions = [
        "", "使用步驟：",
        " 1. 切換到「報名表」分頁。",
        " 2. 每位學生一列，姓名已預先填好。",
        " 3. 在「個人項目1/2/3」和「接力」欄位點下拉選單挑項目，不用打字。",
        " 4. 只需填寫要報名的欄位，不參賽的欄位留空即可。",
        " 5. 儲存檔案，交回體育組或執行下列指令匯入：",
        "        python import_data.py regs 你的檔名.xlsx",
        "",
        "注意事項：",
        " • 系統會在匯入時自動檢核規則，不合法的項目會被拒絕並列出原因。",
        " • 只能選擇符合性別與年級限制的項目（例：五年級不能選限六年級的項目）。",
        " • 「項目清單」和「學生清單」分頁供對照參考，勿修改內容。",
    ]
    for i, t in enumerate(instructions, 4):
        c = ws1.cell(row=i, column=1, value=t)
        c.font = NORMAL_FONT
        if t.startswith(" •") or t.startswith("  "):
            c.fill = INSTR_FILL

    ws_stu = wb.create_sheet("學生清單")
    _apply_header(ws_stu, ["學號", "姓名", "年級", "班級", "座號", "性別", "號碼布"])
    for r, s in enumerate(students, 2):
        for c, v in enumerate(s, 1):
            ws_stu.cell(row=r, column=c, value=v).font = NORMAL_FONT
    for c in range(1, 8):
        ws_stu.column_dimensions[get_column_letter(c)].width = 10

    ws_evt = wb.create_sheet("項目清單")
    _apply_header(ws_evt, ["代碼", "項目名稱", "類別", "性別", "年級限制"])
    for r, e in enumerate(events, 2):
        for c, v in enumerate(e, 1):
            ws_evt.cell(row=r, column=c, value=v).font = NORMAL_FONT
    ws_evt.column_dimensions["A"].width = 12
    ws_evt.column_dimensions["B"].width = 22
    for c in range(3, 6):
        ws_evt.column_dimensions[get_column_letter(c)].width = 10

    ws = wb.create_sheet("報名表", 0)
    headers = ["學號", "姓名", "年級", "班級", "性別",
               "個人項目1", "個人項目2", "個人項目3", "接力", "備註"]
    _apply_header(ws, headers)
    for r, s in enumerate(students, 2):
        sid, name, g, c, seat, gender, bib = s
        for col, v in enumerate([sid, name, g, c, gender], 1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = NORMAL_FONT
            cell.alignment = CENTER
            cell.border = BORDER
            if col <= 5:
                cell.fill = PatternFill("solid", fgColor="F2F2F2")
        for col in range(6, 11):
            ws.cell(row=r, column=col, value="").border = BORDER
    widths = [10, 12, 6, 6, 6, 22, 22, 22, 22, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "F2"

    n_events = len(events)
    n_stu = len(students)

    def _codes_for(category, gender_filter):
        rows = [(i, e) for i, e in enumerate(events, 2) if e[2] == category
                and (e[3] == gender_filter or e[3] == "MIX")]
        return rows

    # 個人項目 (track/field) 下拉：依性別列出全部符合的（年級不好篩，讓匯入時檢核）
    for r in range(2, n_stu + 2):
        stu_gender = ws.cell(row=r, column=5).value
        # 個人項目下拉範圍：跑到項目清單找 track+field 且性別相符
        formulas = []
        for i, e in enumerate(events, 2):
            if e[2] in ("track", "field") and (e[3] == stu_gender or e[3] == "MIX"):
                formulas.append(e[0])
        if formulas:
            dv = DataValidation(type="list", formula1='"' + ",".join(formulas) + '"',
                                allow_blank=True)
            dv.error = "請從下拉選單選擇"
            dv.errorTitle = "無效項目"
            ws.add_data_validation(dv)
            for col in (6, 7, 8):
                dv.add(ws.cell(row=r, column=col))
        # 接力下拉
        relay_codes = [e[0] for e in events if e[2] == "relay"
                       and (e[3] == stu_gender or e[3] == "MIX")]
        if relay_codes:
            dv2 = DataValidation(type="list", formula1='"' + ",".join(relay_codes) + '"',
                                 allow_blank=True)
            ws.add_data_validation(dv2)
            dv2.add(ws.cell(row=r, column=9))

    filename = out_name or f"報名表_{grade_filter or 'all'}{'-' + str(class_filter) if class_filter else ''}_{datetime.now():%Y%m%d}.xlsx"
    out_path = TEMPLATE_DIR / filename
    wb.save(out_path)
    print(f"[OK] 報名表已產出：{out_path}  ({n_stu} 位學生)")
    return out_path


def make_all_classes():
    """為每個班級各產一份"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT grade, class_no FROM students ORDER BY grade, class_no")
    classes = cur.fetchall()
    conn.close()
    for g, c in classes:
        make_form(grade_filter=g, class_filter=c,
                  out_name=f"報名表_{g}年{c}班.xlsx")


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "each":
        make_all_classes()
    else:
        make_form()
        print("\n提示：加參數 'each' 可為每個班級各產一份")
        print("      python make_registration_form.py each")
